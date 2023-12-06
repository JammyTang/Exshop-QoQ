import pandas as pd
import glob
import os
import warnings
warnings.filterwarnings('ignore')
import csv
import openpyxl



# 資料清理
# 授權 Colab 與你的 google 雲端硬碟帳戶連結
from google.colab import drive
drive.mount('/content/drive')


# 本季
# 設定檔案路徑，並讀取資料（官網）
your_path = "/content/drive/MyDrive/exshop/QoQ/raw_data/new_shopline.xlsx"
shopline = pd.read_excel(f"{your_path}")

# 新增銷售管道欄位-官網
shopline.insert(2, "銷售管道", "官網", True)

path_shopline = "/content/drive/MyDrive/exshop/QoQ/data/"
shopline.to_csv(f"{path_shopline}new_shopline.csv", index = False) #index: 列標籤是否要輸出

# 設定檔案路徑，並讀取資料（蝦皮）
your_path = "/content/drive/MyDrive/exshop/QoQ/raw_data/new_shopee.xlsx"
shopee = pd.read_excel(f"{your_path}")

# 新增銷售管道欄位-蝦皮
shopee.insert(2, "銷售管道", "蝦皮", True)

path_shopee = "/content/drive/MyDrive/exshop/QoQ/data/"
shopee.to_csv(f"{path_shopee}new_shopee.csv", index = False) #index: 列標籤是否要輸出


# 設定檔案路徑，並讀取資料（門市）
your_path = "/content/drive/MyDrive/exshop/QoQ/raw_data/new_store.xlsx"
store = pd.read_excel(f"{your_path}")

# 新增銷售管道欄位-門市
store.insert(2, "銷售管道", "門市", True)

path_store = "/content/drive/MyDrive/exshop/QoQ/data/"
store.to_csv(f"{path_store}new_store.csv", index = False) #index: 列標籤是否要輸出


# 設定檔案路徑，並讀取資料（B2B）
your_path = "/content/drive/MyDrive/exshop/QoQ/raw_data/new_b2b.xlsx"
b2b = pd.read_excel(f"{your_path}")

# 新增銷售管道欄位-B2B
b2b.insert(2, "銷售管道", "B2B", True)

path_b2b = "/content/drive/MyDrive/exshop/QoQ/data/"
b2b.to_csv(f"{path_b2b}new_b2b.csv", index = False) #index: 列標籤是否要輸出


# 合併 data 資料夾內的所有 csv 檔
files_joined = os.path.join("/content/drive/MyDrive/exshop/QoQ/data/", "new*.csv")

# 回傳合併後的檔案
list_files = glob.glob(files_joined)

new_all = pd.concat(map(pd.read_csv, list_files), ignore_index = True)



# 刪除不必要欄位
new = new_all.drop(["sales no",
                "customer",
                "consignee",
                "invoice no",
                "invoice date",
                "category no",
                "subCategoryNo",
                "product no",
                "price",
                "tax",
                'equal'], axis = 1
               )


# columns: 要修改的原始欄位名稱與新欄位名稱對應, inplace: 是否直接修改資料表
new.rename(columns = {'ship date': '日期',
                       'product::category':'品類',
                       'product::subcategory':'品牌',
                       'product name':'商品名稱',
                       'quantity':'銷售數量',
                       'price_tax':'單品銷售金額',
                       'total':'銷售金額'},inplace = True
            )



# 新增月份欄位
for i in new.index:
  s = new.loc[i, '日期'][5:7]
  new.loc[i, '月份'] = s

# 將月份欄位移到最前面
col = new.pop('月份')
new.insert(loc = 0, column = '月份', value = col)

# 加入第一列 2023 Q3
q_new = input('本季(ex.2023 Q3)：')
new.insert(0, "銷售季度", f"{q_new}", True)


# 輸出合併後的檔案
your_path = "/content/drive/MyDrive/exshop/QoQ/data/"
new.to_csv(f"{your_path}new.csv", index = False) #index: 列標籤是否要輸出




# 前一季
# 設定檔案路徑，並讀取資料（官網）
your_path = "/content/drive/MyDrive/exshop/QoQ/raw_data/old_shopline.xlsx"
shopline = pd.read_excel(f"{your_path}")

# 新增銷售管道欄位-官網
shopline.insert(2, "銷售管道", "官網", True)

path_shopline = "/content/drive/MyDrive/exshop/QoQ/data/"
shopline.to_csv(f"{path_shopline}old_shopline.csv", index = False) #index: 列標籤是否要輸出

# 設定檔案路徑，並讀取資料（蝦皮）
your_path = "/content/drive/MyDrive/exshop/QoQ/raw_data/old_shopee.xlsx"
shopee = pd.read_excel(f"{your_path}")

# 新增銷售管道欄位-蝦皮
shopee.insert(2, "銷售管道", "蝦皮", True)

path_shopee = "/content/drive/MyDrive/exshop/QoQ/data/"
shopee.to_csv(f"{path_shopee}old_shopee.csv", index = False) #index: 列標籤是否要輸出


# 設定檔案路徑，並讀取資料（門市）
your_path = "/content/drive/MyDrive/exshop/QoQ/raw_data/old_store.xlsx"
store = pd.read_excel(f"{your_path}")

# 新增銷售管道欄位-門市
store.insert(2, "銷售管道", "門市", True)

path_store = "/content/drive/MyDrive/exshop/QoQ/data/"
store.to_csv(f"{path_store}old_store.csv", index = False) #index: 列標籤是否要輸出


# 設定檔案路徑，並讀取資料（B2B）
your_path = "/content/drive/MyDrive/exshop/QoQ/raw_data/old_b2b.xlsx"
b2b = pd.read_excel(f"{your_path}")

# 新增銷售管道欄位-B2B
b2b.insert(2, "銷售管道", "B2B", True)

path_b2b = "/content/drive/MyDrive/exshop/QoQ/data/"
b2b.to_csv(f"{path_b2b}old_b2b.csv", index = False) #index: 列標籤是否要輸出


# 合併 data 資料夾內的所有 csv 檔
files_joined = os.path.join("/content/drive/MyDrive/exshop/QoQ/data", "old*.csv")

# 回傳合併後的檔案
list_files = glob.glob(files_joined)

old_all = pd.concat(map(pd.read_csv, list_files), ignore_index = True)



# 刪除不必要欄位
old = old_all.drop(["sales no",
                "customer",
                "consignee",
                "invoice no",
                "invoice date",
                "category no",
                "subCategoryNo",
                "product no",
                "price",
                "tax",
                'equal'], axis = 1
               )


# columns: 要修改的原始欄位名稱與新欄位名稱對應, inplace: 是否直接修改資料表
old.rename(columns = {'ship date': '日期',
                       'product::category':'品類',
                       'product::subcategory':'品牌',
                       'product name':'商品名稱',
                       'quantity':'銷售數量',
                       'price_tax':'單品銷售金額',
                       'total':'銷售金額'},inplace = True
            )



# 新增月份欄位
for i in old.index:
  s = old.loc[i, '日期'][5:7]
  old.loc[i, '月份'] = s

# 將月份欄位移到最前面
col = old.pop('月份')
old.insert(loc = 0, column = '月份', value = col)

# 加入第一列 銷售季度
q_old = input('上季(ex.2022 Q3)：')
old.insert(0, "銷售季度", f"{q_old}", True)


# 輸出合併後的檔案
your_path = "/content/drive/MyDrive/exshop/QoQ/data/"
old.to_csv(f"{your_path}old.csv", index = False) #index: 列標籤是否要輸出




# 合併為 xlsx 檔
os.chdir('/content/drive/MyDrive/exshop/QoQ/data')  # Colab 換路徑使用

csvfile_old = open('old.csv')     # 開啟 CSV 檔案
raw_data_old = csv.reader(csvfile_old)     # 讀取 CSV 檔案
data_old = list(raw_data_old)              # 轉換成二維串列
wb = openpyxl.Workbook()
sheet_old = wb.create_sheet(f"{q_old}")     # 建立空白的工作表
for i in data_old:
    sheet_old.append(i)                # 逐筆添加到最後一列



csvfile_new = open('new.csv')
raw_data_new = csv.reader(csvfile_new)
data_new = list(raw_data_new)
sheet_new = wb.create_sheet(f"{q_new}")
for i in data_new:
    sheet_new.append(i)



# 刪除不需要的工作表
wb.remove(wb['Sheet'])


# 儲存工作表
file = input('季報名稱：')
wb.save(f"{file}.xlsx")


# 將月份、銷售數量、單品銷售金額、銷售金額欄位轉換為數字格式
df1 = pd.read_excel(f"/content/drive/MyDrive/exshop/QoQ/data/{file}.xlsx", sheet_name = f"{q_new}")
df1['月份'] = pd.to_numeric(df1['月份'], errors='coerce')
df1['銷售數量'] = pd.to_numeric(df1['銷售數量'], errors='coerce')
df1['單品銷售金額'] = pd.to_numeric(df1['單品銷售金額'], errors='coerce')
df1['銷售金額'] = pd.to_numeric(df1['銷售金額'], errors='coerce')

df2 = pd.read_excel(f"/content/drive/MyDrive/exshop/QoQ/data/{file}.xlsx", sheet_name = f"{q_old}")
df2['月份'] = pd.to_numeric(df2['月份'], errors='coerce')
df2['銷售數量'] = pd.to_numeric(df2['銷售數量'], errors='coerce')
df2['單品銷售金額'] = pd.to_numeric(df2['單品銷售金額'], errors='coerce')
df2['銷售金額'] = pd.to_numeric(df2['銷售金額'], errors='coerce')


# 將更改後的 DataFrame 保存回 Excel 文件
with pd.ExcelWriter(f"/content/drive/MyDrive/exshop/QoQ/data/{file}.xlsx") as writer:
    df1.to_excel(writer, sheet_name= f"{q_new}", index=False)
    df2.to_excel(writer, sheet_name= f"{q_old}", index=False)
